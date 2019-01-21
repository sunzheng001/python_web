# -*- coding: utf-8 -*-
# Author: sunzheng
# Time: 2018/12/20
#

from  openpyxl import load_workbook

opne_excel=load_workbook("C:\\Users\\Administrator\\Desktop\\量体计划单模板(1).xlsx")
sheet_name=opne_excel["Sheet1"]
test=[]
for itme  in range(2,4):
    sub_data={}
    sub_data["姓名"]=sheet_name.cell(itme,1).value
    sub_data["手机号码"] = sheet_name.cell(itme, 2).value
    sub_data["款号"] = sheet_name.cell(itme, 3).value
    test.append(sub_data)
print(test)
    # for i in range(1,4):
    #     sub_data.append(sheet_name.cell(itme,i).value)
    # test.append(sub_data)
print(test)
# do_excel=sheet_name.cell(1,2).value
# print(do_excel)

#
#
#
# from openpyxl import load_workbook
#
# worksheet =load_workbook("D:\\testdata_TestCase.xlsx")#打开excel文件
# # sheet_name=worksheet.get_sheet_by_name("sheet1")
# sheet_test=worksheet["testdata"]
# result01=sheet_test.cell(1,1).value
# print(result01)
#
# # sheet_test.cell(1,2).value="sunzheng"#修改excel文件数据
# # worksheet.save("D:\\工作簿1(1).xlsx")#修改后保存数据